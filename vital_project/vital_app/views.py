from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from .models import Vehicle
from django.db.models import Q
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.http import HttpResponse
from .models import Vehicle
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from . import views  


def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)

            # Check if the user is an admin (staff member)
            if user.is_staff:
                return redirect('summary')
            else:
                return redirect('vehicle')

    else:
        form = AuthenticationForm()
    return render(request, 'vital_app/registration/login.html', {'form': form})

@staff_member_required
def summary_view(request):
    # Get the search field and query from the GET request
    search_field = request.GET.get('search_field', 'all')
    search_query = request.GET.get('search', '')

    # Define a list of model fields to search through
    search_fields = [
        'fleet_no',
        'cost_centre',
        'reg_no',
        'category',
        'group',
        'body_type',
        'area_available',
        # Add fields for other criteria
    ]

    # If the selected field is "All Fields," construct a query that searches in all fields
    if search_field == 'all':
        query = Vehicle.objects.none()  # Start with an empty queryset
        for field in search_fields:
            query |= Vehicle.objects.filter(**{f"{field}__icontains": search_query})
    else:
        # If a specific field is selected, filter the queryset based on that field
        query = Vehicle.objects.filter(**{f"{search_field}__icontains": search_query})

    context = {
        'vehicles': query,  # Display search results based on the selected field
        'search_field': search_field,
        'search_query': search_query,
    }

    return render(request, 'vital_app/registration/summary.html', context)


@login_required
def vehicle_view(request):
    # Get the search field and query from the GET request
    search_field = request.GET.get('search_field', 'all')
    search_query = request.GET.get('search', '')

    # Define a list of model fields to search through
    search_fields = [
        'fleet_no',
        'cost_centre',
        'reg_no',
        'category',
        'group',
        'body_type',
        'area_available',
        # Add fields for other criteria
    ]

    # If the selected field is "All Fields," construct a query that searches in all fields
    if search_field == 'all':
        query = Vehicle.objects.none()  # Start with an empty queryset
        for field in search_fields:
            query |= Vehicle.objects.filter(**{f"{field}__icontains": search_query})
    else:
        # If a specific field is selected, filter the queryset based on that field
        query = Vehicle.objects.filter(**{f"{search_field}__icontains": search_query})

    context = {
        'vehicles': query,  # Display search results based on the selected field
        'search_field': search_field,
        'search_query': search_query,
    }
    return render(request, 'vital_app/registration/vehicle.html', context)


@csrf_exempt
def update_status(request, vehicleId, status):
    if request.method == 'POST':
        print(f"Attempting to update vehicle: {vehicleId} with status: {status}")
        try:
            vehicle = Vehicle.objects.get(reg_no=vehicleId)
            
            # If the status is set to 'workshop', record the current time
            if status == "workshop" and vehicle.status != "workshop":
                vehicle.workshop_entry_time = timezone.now()
            # If the status changes from 'workshop' to something else, nullify the time
            elif status != "workshop" and vehicle.status == "workshop":
                vehicle.workshop_entry_time = None

            vehicle.status = status
            vehicle.save()
            print(f"Updated vehicle {vehicleId} to status {status}")
            return JsonResponse({'success': True})
        except Vehicle.DoesNotExist:
            print(f"Vehicle with ID {vehicleId} does not exist.")
            return JsonResponse({'success': False, 'error': 'Vehicle does not exist'})
        except Exception as e:
            print(f"Error updating vehicle: {e}")
            return JsonResponse({'success': False, 'error': str(e)})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

# views.py
def graphs(request):
    # Fetch the count of trucks based on their status
    available_count = Vehicle.objects.filter(status="available").count()
    hired_out_count = Vehicle.objects.filter(status="hired_out").count()
    workshop_count = Vehicle.objects.filter(status="workshop").count()

    context = {
        'available_count': available_count,
        'hired_out_count': hired_out_count,
        'workshop_count': workshop_count
    }

    return render(request, 'vital_app/registration/graphs.html', context)

def download_report(request, status):
    # Fetch vehicles based on the status
    vehicles = Vehicle.objects.filter(status=status)

    # Create a new workbook and add a worksheet.
    wb = Workbook()
    ws = wb.active
    ws.title = f"{status.capitalize()} Vehicles"

    # Define headers
    headers = [
        'Fleet Number', 'Cost Centre', 'Registration Number', 
        'Category', 'Group', 'Body Type', 'Status', 'Area Available',
        'Workshop Duration'
    ]

    # Add headers to the worksheet
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws['{}1'.format(col_letter)] = header
        ws.column_dimensions[col_letter].width = 15

    # Populate the worksheet with data
    for row_num, vehicle in enumerate(vehicles, 2):
        ws.cell(row=row_num, column=1, value=vehicle.fleet_no)
        ws.cell(row=row_num, column=2, value=vehicle.cost_centre)
        ws.cell(row=row_num, column=3, value=vehicle.reg_no)
        ws.cell(row=row_num, column=4, value=vehicle.category)
        ws.cell(row=row_num, column=5, value=vehicle.group)
        ws.cell(row=row_num, column=6, value=vehicle.body_type)
        ws.cell(row=row_num, column=7, value=vehicle.status)
        ws.cell(row=row_num, column=8, value=vehicle.area_available)
        ws.cell(row=row_num, column=9, value=vehicle.workshop_duration)

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{status}_vehicles_report.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response

# views.py
from django.shortcuts import render

def report_view(request):
    # Your logic to retrieve and render the content for the report page
       return render(request, 'vital_app/registration/reports.html')


  
